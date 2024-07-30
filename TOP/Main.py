from openpyxl import Workbook
import pandas as pd
import numpy as np
from VND import VND
from openpyxl import load_workbook
import copy
from GRASP_GA import GRASP
from GA import GA
from GA_VND import GA_VND

def calcular_long_total(ruta, edges):
    longitud = 0
    i = 0
    while i < len(ruta)-1:
        longitud = longitud + edges[(f"{ruta[i]}", f"{ruta[i+1]}")]
        i+=1
    return longitud

def calcular_peso_total(ruta, weights):
    peso = 0
    i = 0
    while i < len(ruta)-1:
        peso= peso + weights[f"{ruta[i]}"]
        i+=1
    return peso

def crear_excel(id, rutas, tiempo, weights, edges, criterio):
    ruta_modificada = copy.deepcopy(rutas)
    nombre_archivo = f'TOP_Castrillon_correa_{criterio}.xlsx'
    wb = load_workbook(nombre_archivo)
    ws = wb.create_sheet(title=f"TOP{id}")
    longitud_rutas = []
    peso_rutas = []
    for w in ruta_modificada:
        longitud_rutas.append(calcular_long_total(w, edges))
        peso_rutas.append(calcular_peso_total(w, weights))
    for ruta in ruta_modificada:
        for pos in range(len(ruta)):
            nodo = int(ruta[pos]) + 1
            ruta[pos] = str(nodo)
    peso_total = 0
    for h in peso_rutas:
        peso_total = peso_total + h
    fila = 0
    for i in ruta_modificada:
        fila += 1
        ws[f'A{fila}'] = len(i)
        for z, valor in enumerate(i, start=2):
            ws.cell(row=fila, column=z, value=valor)    
        ws.cell(row=fila, column=len(i) + 2, value=longitud_rutas[fila-1])
        ws.cell(row=fila, column=len(i) + 3, value=peso_rutas[fila-1])
    fila += 1
    ws[f'A{fila}'] = peso_total
    ws[f'B{fila}'] = tiempo
    wb.save(nombre_archivo)

def read_fsspsc(id):
    # Construir el nombre del archivo basado en el identificador dado
    file_path = f'TOP{id}.txt'
    time_limit = f'TimeLimit.xlsx'

    data = pd.read_excel(time_limit, header=None)
    
    nuevos_titulos = ['TOP', 'Time']  # Asegúrate de que la longitud de esta lista coincida con el número de columnas
    data.columns = nuevos_titulos
    
    time = data['Time'][id-1]
    
    # Leer el archivo
    with open(file_path, 'r') as file:
        lines = file.readlines()

    # Procesar las líneas para obtener 'n', 'm', 'L'
    n = int(lines[0].strip())
    m = int(lines[1].strip())
    L = float(lines[2].strip())

    # Procesar las filas restantes para 'dur'
    dur = np.array([list(map(float, line.strip().split())) for line in lines[3:]])

    return n, m, L, dur, time

def convert_to_hybrid_representation(grasp_population, num_nodes):
    hybrid_population = []
    for solution in grasp_population:
        hybrid_solution = []
        for route in solution:
            binary_route = [0] * num_nodes
            for node in route:
                binary_route[int(node)] = 1  # Convertir el string a entero
            hybrid_solution.append((route, binary_route))
        hybrid_population.append(hybrid_solution)
    return hybrid_population


nombres_archivos = ['TOP_Castrillon_correa_GA_VND1.xlsx', 'TOP_Castrillon_correa_GA1.xlsx']


for nombre in nombres_archivos:
    wb = Workbook()    
    wb.save(nombre)

nsol= 50 # Number of solutions
num_children = 25
mutation_prob = 0.01

busquedas = ['Intercambio', 'Invertir', 'Cambiar_otra_ruta']
for id in range(1, 22):
    print(id)
    n,m,L,dur, time_limit = read_fsspsc(id)
    population, t0, edges, weights=GRASP(n,m,L,dur,id,nsol)
    hybrid_population = convert_to_hybrid_representation(population, n)
    GA_solution, Aptitud_GA_solution, tGA = GA(nsol, hybrid_population, time_limit, num_children, mutation_prob, t0, L, edges, weights, n)
    GA_solution_VND, Aptitud_GA_solution_VND, tGA_VND = GA_VND(nsol, hybrid_population, time_limit, num_children, mutation_prob, t0, L, edges, weights, n, busquedas)
    final_route = []
    longGA = 0
    for route_in, binary_route in GA_solution:
        final_route.append(route_in)
        longGA += calcular_long_total(route_in, edges)
    print("GA:\t", Aptitud_GA_solution, "\t", tGA, "\t", longGA)
    crear_excel(id, final_route, tGA, weights, edges, 'GA1')
    final_route_VND = []
    longGA_VND = 0
    for route_in, binary_route in GA_solution_VND:
        final_route_VND.append(route_in)
        longGA_VND += calcular_long_total(route_in, edges)
    print("GA_VND:\t", Aptitud_GA_solution_VND, "\t", tGA_VND, "\t", longGA_VND)
    crear_excel(id, final_route_VND, tGA_VND, weights, edges, 'GA_VND1')
    

for nombre in nombres_archivos:
   wb = load_workbook(nombre)
   wb.remove(wb['Sheet'])
   wb.save(nombre)
